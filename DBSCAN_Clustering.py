import os

os.environ['OPENBLAS_NUM_THREADS'] = '1'

import pandas as pd
import matplotlib.pyplot as plt
import dask.dataframe as dd
from datetime import datetime
import numpy as np
from sklearn.preprocessing import LabelEncoder, OneHotEncoder
from sklearn.impute import SimpleImputer
from sklearn.impute import SimpleImputer
from sklearn.metrics import accuracy_score
from sklearn.metrics import f1_score
from sklearn.metrics import recall_score
from sklearn.metrics import precision_score
from sklearn.metrics import PrecisionRecallDisplay
from sklearn.cluster import DBSCAN
from sklearn import metrics
from matplotlib.widgets import Slider
import winsound
import xlsxwriter
from openpyxl import load_workbook
import os.path
import csv

def pd_set_df_view_options(max_rows=1000, max_columns=350, display_width=320):

    # Show more than 10 or 20 rows when a dataframe comes back.
    pd.set_option('display.max_rows', max_rows)
    # Columns displayed in debug view
    pd.set_option('display.max_columns', max_columns)

    pd.set_option('display.width', display_width)

# run
pd_set_df_view_options(max_rows=1000, max_columns=350, display_width=320)


capture_date_start = 1665100800000000
capture_date_end = 1665187200000000
df = pd.read_csv("C:\\CICIoT2023\\csv\\benign\\BenignTraffic_Merged_01percent.csv")
column_name = 'Timestamp'  # Replace with your actual column name
df.columns = df.columns.str.strip()
df = df.replace([np.inf,-np.inf], np.nan)


if 'ts' in df.columns:
    df.insert(df.columns.get_loc("ts") + 1, column_name,(df['ts']*1000000)-capture_date_start , True)
    df[column_name] = df[column_name].astype('Int64')
    df = df.drop(columns=['ts'])
if 'IAT' in df.columns:
    df = df.drop(columns=['IAT'])
if 'max_duration' in df.columns:
    df = df.drop(columns=['max_duration'])
if 'min_duration' in df.columns:
    df = df.drop(columns=['min_duration'])
if 'average_duration' in df.columns:
    df = df.drop(columns=['average_duration'])



min_value = df[column_name].min()
max_value = df[column_name].max()

timedist = (max_value-min_value)/8

filtered_df = df
# filtered_df = df[
#     ((df[column_name] > min_value+0*timedist) & (df[column_name] <= min_value+1*timedist))|
#     ((df[column_name] > min_value+2*timedist) & (df[column_name] <= min_value+3*timedist))|
#     ((df[column_name] > min_value+4*timedist) & (df[column_name] <= min_value+5*timedist))|
#     ((df[column_name] > min_value+6*timedist) & (df[column_name] <= min_value+7*timedist))
#       ]

X = filtered_df.loc[:, :].values
labelencoder_X = LabelEncoder()
X[:, filtered_df.columns.get_loc("Protocol_name")] = labelencoder_X.fit_transform(X[:, filtered_df.columns.get_loc("Protocol_name")])


imputer = SimpleImputer(missing_values = np.nan, strategy = "mean")
imputer = SimpleImputer.fit(imputer,X[:,:])
X[:, :] = imputer.transform(X[:, :])
X[:, :] = imputer.transform(X[:, :])


print("Start dbscan")

# from sklearn.neighbors import NearestNeighbors
# from matplotlib import pyplot as plt
#
# neighbors = NearestNeighbors(n_neighbors=120)
# neighbors_fit = neighbors.fit(X)
# distances, indices = neighbors_fit.kneighbors(X)
# distances = np.sort(distances, axis=0)
# distances = distances[:,1]
# plt.xlim(2400, 6000)
# plt.plot(distances)
# plt.grid(True, linestyle='-', alpha=0.1)
# plt.ylim(0, 100000000000)  # Adjust the y-axis limits as needed
# plt.show()
# exit()
name = 'diagrams5'
if not os.path.exists('C:\\CICIoT2023\\dbscan\\'+name):
    os.makedirs('C:\\CICIoT2023\\dbscan\\'+name)
output_file ='C:\\CICIoT2023\\dbscan\\BENIGNS_DBSCAN_'+name+'.csv'

# if not os.path.exists('C:\\CICIoT2023\\dbscan\\BENIGNS_DBSCAN.xlsx'):
#     workbook = xlsxwriter.Workbook('C:\\CICIoT2023\\dbscan\\BENIGNS_DBSCAN.xlsx')
#     worksheet = workbook.add_worksheet(name="DBSCAN")
#     worksheet.write('A1', 'EPS')
#     worksheet.write('B1', 'MinSample')
#     worksheet.write('C1', 'Cluster')
#     worksheet.write('D1', 'Noise')
#     worksheet.write('E1', 'NoisePercent')
#     workbook.close()

i = 1

for ieps in range(5,31):
    for minsamp in range(1,20):

        eps = ieps*1000000000
        minsample = minsamp*10
        db = DBSCAN(eps=eps, min_samples=minsample).fit(X)

        # core_samples_mask = np.zeros_like(db.labels_, dtype=bool)
        # core_samples_mask[db.core_sample_indices_] = True
        labels = db.labels_

        # Number of clusters in labels, ignoring noise if present.
        n_clusters_ = len(set(labels)) - (1 if -1 in labels else 0)
        n_noise_ = list(labels).count(-1)
        # workbook1 = load_workbook('C:\\CICIoT2023\\dbscan\\BENIGNS_DBSCAN.xlsx')
        # worksheet1 = workbook1.active
        # worksheet1.append([eps,minsample,n_clusters_,n_noise_,(n_noise_ / len(X) * 100)])
        with open(output_file, 'a', newline='') as csvfile:
            csv_writer = csv.writer(csvfile)
            data_row = [[i,len(X),eps,minsample,n_clusters_,n_noise_,(n_noise_ / len(X) * 100)]]
            csv_writer.writerows(data_row)

        if 'unixcluster' in filtered_df.columns:
            filtered_df = filtered_df.drop(columns=['unixcluster'])
        filtered_df['unixcluster'] = labels
        print(list(set(filtered_df['unixcluster'])))
        # for label in labels:
        #     if label == -1:
        #         continue
        #     else:
        #         continue
        plt.xlabel(column_name)
        plt.scatter(filtered_df[column_name], filtered_df['unixcluster'], c="blue", s=1)
        plt.ylabel('Clusters')
        plt.title('Distribution of ' + column_name + ' based on clusters')
        plt.savefig('C:\\CICIoT2023\\dbscan\\'+name+'\\' + str(int(n_noise_ / len(X) * 100)) +'-'+ str(n_clusters_) + '-' + str(eps) + '_' + str(
            minsample) + '.png')
        plt.clf()

        print("(",str(i),"/",str(5000),")")
        print("eps:",eps," minsample:",minsample)
        print("Number of records: %d" % len(X))
        print("Estimated number of clusters: %d" % n_clusters_)
        print("Estimated number of noise points: %d" % n_noise_)
        print("Estimated percent of noise points: %d" % float(n_noise_ / len(X) * 100))

        print("_______________________________________________________")
        # print("Adjusted Mutual Information: %0.3f" % metrics.adjusted_mutual_info_score(y, labels))
        # print("Silhouette Coefficient: %0.3f" % metrics.silhouette_score(X, labels))
        winsound.Beep(440, 500)
        i+=1
        # workbook1.save('C:\\CICIoT2023\\dbscan\\BENIGNS_DBSCAN.xlsx')
        # workbook1.close()

#
# # Plotting the grouped data
# #plt.bar(grouped_data[column_name], grouped_data['Frequency'])
# plt.hist(df[column_name], bins=30, density=True)
# plt.xlabel(column_name)
# plt.ylabel('Frequency')
# plt.title('Grouped Frequency Plot')
# plt.show()
